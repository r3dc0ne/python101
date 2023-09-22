from openpyxl import load_workbook

workbook_boninites = load_workbook(
    filename="G:/My Drive/PycharmProjects/coding101/matplotlip_petro/magmatic/boninites_excel.xlsx"
)
sheet_original = workbook_boninites["original"]

workbook_true_boninites = load_workbook(
    filename="G:/My Drive/PycharmProjects/coding101/matplotlip_petro/magmatic/true_boninites.xlsx"
)
sheet_true_boninites = workbook_true_boninites["true_boninites"]


