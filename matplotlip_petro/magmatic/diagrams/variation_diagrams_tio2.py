import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from helpful_functions import get_column_values

# all boninites
workbook_original = load_workbook(
    filename="G:/My Drive/PycharmProjects/coding101/matplotlip_petro/magmatic/boninites_excel.xlsx"
)
worksheet_original = workbook_original["original"]

list_SiO2_all = get_column_values(worksheet_original, 28)
list_TiO2_all = get_column_values(worksheet_original, 29)

# true boninites
workbook_true_boninites = load_workbook(
    filename="G:/My Drive/PycharmProjects/coding101/matplotlip_petro/magmatic/true_boninites.xlsx"
)
worksheet_true = workbook_true_boninites["true_boninites"]

list_SiO2_true = get_column_values(worksheet_true, 28)
list_TiO2_true = get_column_values(worksheet_true, 29)

# diagram plotter
plt.figure(figsize=(8, 6))

plt.axvline(x=52, color='grey', linestyle='-', linewidth=0.5, zorder=1)
plt.axvline(x=63, color='grey', linestyle='-', linewidth=0.5, zorder=1)
plt.axhline(y=0.5, color='grey', linestyle='-', linewidth=0.5, zorder=1)

plt.scatter(list_SiO2_all, list_TiO2_all, color="blue", s=15, label="GEOROC Boninites", zorder=2)
plt.scatter(list_SiO2_true, list_TiO2_true, color="orange", s=10, label="IUGS Boninites", zorder=3)

plt.xlabel('SiO\u2082 wt.%')
plt.ylabel('TiO\u2082 wt.%')
plt.title('Variation diagram: TiO\u2082 wt.% over SiO\u2082 wt.%')

plt.xlim(30, 80)
plt.ylim(0, 2)

plt.legend()
plt.savefig('ti_silica.png', dpi=300)
plt.show()


# hochgestellt: \xb
# tiefgestellt: \u208
