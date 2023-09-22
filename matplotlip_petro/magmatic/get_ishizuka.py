from openpyxl import load_workbook
from boninites_functions import get_all_rows_with_value, copy_row, delete_rows_by_comparison


workbook_true_boninites = load_workbook(
    filename="G:/My Drive/PycharmProjects/coding101/matplotlip_petro/magmatic/true_boninites.xlsx"
)
sheet_true_boninites = workbook_true_boninites["true_boninites"]
sheet_ishizuka = workbook_true_boninites.create_sheet("ishizuka")

copy_row(sheet_true_boninites, 1, sheet_ishizuka, 1)
get_all_rows_with_value("[24514]", 0, sheet_true_boninites, sheet_ishizuka)

workbook_true_boninites.save(
    filename="G:/My Drive/PycharmProjects/coding101/matplotlip_petro/magmatic/boninites_ishizuka.xlsx")

