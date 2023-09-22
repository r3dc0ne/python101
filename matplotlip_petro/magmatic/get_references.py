from openpyxl import Workbook
from openpyxl import load_workbook

workbook_boninites = load_workbook(
    filename="G:/My Drive/PycharmProjects/coding101/matplotlip_petro/magmatic/boninites_excel.xlsx"
)
sheet_original = workbook_boninites["original"]

workbook_references = Workbook()
sheet_references = workbook_references.create_sheet("references")

# Find the first row with the specific value in column A
start_row = None
for row in sheet_original.iter_rows(min_row=2):
    if row[0].value == 'References:':
        start_row = row[0].row
        break

# Copy the cells from start_row to the last row in Sheet1 to Sheet2
for row in sheet_original.iter_rows(min_row=start_row, values_only=True):
    sheet_references.append(row)

workbook_references.save(
    filename="G:/My Drive/PycharmProjects/coding101/matplotlip_petro/magmatic/boninites_references.xlsx"
)

